import os
from typing import Dict, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PIL import Image as PILImage

# Resolve directories relative to this file inside catalog/scripts
FILE_DIR = os.path.dirname(__file__)
CATALOG_DIR = os.path.normpath(os.path.join(FILE_DIR, ".."))
DOCS_DIR = os.path.join(CATALOG_DIR, "docs")
XLSX_PATH = os.path.join(DOCS_DIR, "defect_catalog.xlsx")
IMAGES_ROOT = os.path.join(DOCS_DIR, "images")

# Defect metadata (edit as needed)
DEFECTS: Dict[str, Dict[str, str]] = {
	
	
}

COLS = ["defect_type", "summary", "typical_causes", "thermal_signature", "example_image"]
# Widen text columns for readability
COL_WIDTHS = {"A": 22, "B": 60, "C": 52, "D": 56, "E": 36}
# Embed constraints (display only)
EMBED_MAX_W = 600
EMBED_MAX_H = 185
ROW_HEIGHT = 145


def ensure_parent_dir(path: str) -> None:
	os.makedirs(os.path.dirname(path), exist_ok=True)


def build():
	wb = Workbook()
	ws = wb.active
	ws.title = "Defect Catalog"
	for col, width in COL_WIDTHS.items():
		ws.column_dimensions[col].width = width

	# Header
	ws.append(COLS)
	for cell in ws[1]:
		cell.font = cell.font.copy(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")

	row = 2
	wrap_top = Alignment(wrap_text=True, vertical="top")
	for defect, meta in DEFECTS.items():
		ws.row_dimensions[row].height = ROW_HEIGHT
		ws.cell(row=row, column=1, value=defect).alignment = wrap_top
		ws.cell(row=row, column=2, value=meta.get("summary", "")).alignment = wrap_top
		ws.cell(row=row, column=3, value=meta.get("typical_causes", "")).alignment = wrap_top
		ws.cell(row=row, column=4, value=meta.get("thermal_signature", "")).alignment = wrap_top

		img_file = os.path.join(IMAGES_ROOT, f"{defect}.jpg")
		if os.path.exists(img_file):
			try:
				ximg = XLImage(img_file)
				# Scale to fit within bounds (preserve aspect)
				try:
					with PILImage.open(img_file) as im:
						w, h = im.size
					scale = min(EMBED_MAX_W / float(w), EMBED_MAX_H / float(h))
					scale = min(scale, 1.0)
					ximg.width = int(w * scale)
					ximg.height = int(h * scale)
				except Exception:
					ximg.width = EMBED_MAX_W
					ximg.height = EMBED_MAX_H
				ws.add_image(ximg, f"E{row}")
			except Exception:
				ws.cell(row=row, column=5, value="(image embed failed)").alignment = wrap_top
		else:
			ws.cell(row=row, column=5, value="(image not found)").alignment = wrap_top
		row += 1

	ensure_parent_dir(XLSX_PATH)
	wb.save(XLSX_PATH)
	print(f"Saved: {XLSX_PATH}")


if __name__ == "__main__":
	build()
